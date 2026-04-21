#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
必榮 進料標籤列印系統
跨電腦通用版本 - 不需要授權碼，任何 Windows 電腦皆可執行
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import sqlite3
import os
import json
import win32print
import win32ui
import win32con
import win32gui
from PIL import Image, ImageDraw, ImageFont
import qrcode
import openpyxl
from datetime import datetime

# ── 設定常數 ──────────────────────────────────────────────────────────────────
APP_TITLE   = "必榮 進料標籤列印系統"
_BASE       = os.path.dirname(os.path.abspath(__file__))
DB_PATH     = os.path.join(_BASE, "FastReport_sqllite.db")
CONFIG_PATH = os.path.join(_BASE, "settings.json")

LABEL_W_MM  = 99.86
LABEL_H_MM  = 59.3
MARGIN_MM   = 2.5
PRINT_DPI   = 203
LABEL_W_PX  = int(LABEL_W_MM / 25.4 * PRINT_DPI)
LABEL_H_PX  = int(LABEL_H_MM / 25.4 * PRINT_DPI)
MARGIN_PX   = int(MARGIN_MM   / 25.4 * PRINT_DPI)

FONT_PATH      = "C:/Windows/Fonts/msjh.ttc"
FONT_BOLD_PATH = "C:/Windows/Fonts/msjhbd.ttc"

COLUMNS = [
    ("序號", 70), ("供應商名稱", 80), ("訂單編號", 90),
    ("材質", 50), ("尺寸", 50), ("批號", 60), ("特殊", 60),
    ("長度", 60), ("數量", 50), ("製造編號/爐號", 120), ("進貨日期", 90),
]


# ── 設定檔（視窗記憶）────────────────────────────────────────────────────────

def load_config() -> dict:
    try:
        with open(CONFIG_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


def save_config(cfg: dict):
    try:
        with open(CONFIG_PATH, "w", encoding="utf-8") as f:
            json.dump(cfg, f, ensure_ascii=False, indent=2)
    except Exception:
        pass


# ── 資料庫 ────────────────────────────────────────────────────────────────────

def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS MYTABLE (
            INDX INT, SN INT, QTY INT, SEL NUM,
            序號 TEXT, 供應商名稱 TEXT, 訂單編號 TEXT, 材質 TEXT,
            尺寸 TEXT, 批號 TEXT, 特殊 TEXT, 長度 TEXT,
            數量 INT, 重量 TEXT, 不良支數 TEXT, 樣品量 TEXT,
            檢驗尺寸 TEXT, 檢驗外觀 TEXT, 檢驗材質 TEXT, 判定 TEXT,
            製造編號爐號 TEXT, 進貨日期 TEXT, 備註 TEXT,
            PKGQTY INT, SNN INT
        )
    """)
    # 列印記錄表
    cur.execute("""
        CREATE TABLE IF NOT EXISTS PRINT_LOG (
            ID        INTEGER PRIMARY KEY AUTOINCREMENT,
            SN        INTEGER NOT NULL,
            PRINTED_AT TEXT    NOT NULL,
            PRINTER   TEXT,
            LABEL_NUMS TEXT
        )
    """)
    conn.commit()
    conn.close()


def load_table_data():
    conn = get_db()
    cur = conn.cursor()
    try:
        cur.execute("""
            SELECT 序號, 供應商名稱, 訂單編號, 材質, 尺寸, 批號, 特殊,
                   長度, 數量, "製造編號/爐號", 進貨日期, PKGQTY, SNN, SN
            FROM MYTABLE ORDER BY SN
        """)
        return cur.fetchall()
    except Exception:
        return []
    finally:
        conn.close()


def get_printed_sns() -> set:
    """回傳已列印過的 SN 集合"""
    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute("SELECT DISTINCT SN FROM PRINT_LOG")
        return {row[0] for row in cur.fetchall()}
    except Exception:
        return set()
    finally:
        conn.close()


def log_print(sn: int, printer: str, label_nums: list):
    """寫入列印記錄"""
    conn = get_db()
    try:
        conn.execute(
            "INSERT INTO PRINT_LOG (SN, PRINTED_AT, PRINTER, LABEL_NUMS) VALUES (?,?,?,?)",
            (sn, datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
             printer, json.dumps(label_nums))
        )
        conn.commit()
    except Exception:
        pass
    finally:
        conn.close()


def get_print_history(sn: int) -> list:
    """取得某筆資料的所有列印記錄"""
    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute(
            "SELECT PRINTED_AT, PRINTER, LABEL_NUMS FROM PRINT_LOG WHERE SN=? ORDER BY ID DESC",
            (sn,)
        )
        return cur.fetchall()
    except Exception:
        return []
    finally:
        conn.close()


def import_excel_to_db(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]

    conn = get_db()
    cur = conn.cursor()
    cur.execute("DELETE FROM MYTABLE")
    cur.execute("DELETE FROM PRINT_LOG")  # 清除舊列印記錄，避免 SN 重複誤判已印

    sn = 1
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue

        def get(col_name):
            try:
                idx = headers.index(col_name)
                v = row[idx]
                if v is None:
                    return ""
                if isinstance(v, datetime):
                    return v.strftime("%Y/%m/%d")
                return str(v)
            except (ValueError, IndexError):
                return ""

        cur.execute("""
            INSERT INTO MYTABLE
            (SN, QTY, SEL, 序號, 供應商名稱, 訂單編號, 材質, 尺寸, 批號, 特殊,
             長度, 數量, 重量, 不良支數, 樣品量, 檢驗尺寸, 檢驗外觀, 檢驗材質,
             判定, "製造編號/爐號", 進貨日期, 備註, PKGQTY, SNN)
            VALUES (?,1,1,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,1,NULL)
        """, (
            sn,
            get("序號"), get("供應商名稱"), get("訂單編號"), get("材質"),
            get("尺寸"), get("批號"), get("特殊"), get("長度"), get("數量"),
            get("重量"), get("不良支數"), get("樣品量"), get("檢驗尺寸"),
            get("檢驗外觀"), get("檢驗材質"), get("判定"), get("製造編號/爐號"),
            get("進貨日期"), get("備註"),
        ))
        sn += 1

    conn.commit()
    conn.close()
    return sn - 1


# ── 標籤圖像產生 ───────────────────────────────────────────────────────────────

def _load_font(path, size):
    try:
        return ImageFont.truetype(path, size)
    except Exception:
        return ImageFont.load_default()


def _draw_cell(draw, x1, y1, x2, y2, text, font, h_align="center", pad=6):
    """在格子內置中（或靠左）繪製文字，自動截斷過長文字"""
    cell_w = x2 - x1
    cell_h = y2 - y1
    while text:
        bbox = font.getbbox(text)
        if bbox[2] - bbox[0] <= cell_w - pad * 2:
            break
        text = text[:-1]

    bbox = font.getbbox(text)
    tw = bbox[2] - bbox[0]
    th = bbox[3] - bbox[1]
    tx = x1 + (cell_w - tw) // 2 if h_align == "center" else x1 + pad
    ty = y1 + (cell_h - th) // 2 - bbox[1]
    draw.text((tx, ty), text, fill="black", font=font)


def make_label_image(record, pkg_no=1, pkg_total=1):
    """產生一張標籤的 PIL Image，四周保留 MARGIN_PX 留白"""
    W, H = LABEL_W_PX, LABEL_H_PX
    M = MARGIN_PX

    img  = Image.new("RGB", (W, H), "white")
    draw = ImageDraw.Draw(img)

    font_lbl  = _load_font(FONT_BOLD_PATH, 36)
    font_data = _load_font(FONT_BOLD_PATH, 36)
    font_bot  = _load_font(FONT_BOLD_PATH, 32)

    CX, CY = M, M
    CW, CH = W - 2*M, H - 2*M

    BOT_H  = int(CH * 0.09)
    MAIN_H = CH - BOT_H
    ROW_H  = MAIN_H // 8
    MAIN_H = ROW_H * 8

    LBL_W = int(CW * 0.331)
    QR_W  = int(CW * 0.260)
    X_LBL  = CX
    X_DATA = CX + LBL_W
    X_QR   = CX + CW - QR_W
    X_END  = CX + CW
    QR_ROWS = 4

    # QR Code
    qr_text = (
        f"供應商 : {record.get('供應商名稱','')}\n"
        f"進貨日期 : {record.get('進貨日期','')}\n"
        f"材質/特殊 : {record.get('材質','')}{record.get('特殊','')}\n"
        f"尺寸 : {record.get('尺寸','')}\n"
        f"長度 : {record.get('長度','')}\n"
        f"批號 : {record.get('批號','')}\n"
        f"數量 : {pkg_no}/{pkg_total}\n"
        f"製造編號/爐號 : {record.get('製造編號/爐號','')}\n"
        f"ERP序號 : {record.get('序號','')}\n"
        f"訂單編號 : {record.get('訂單編號','')}"
    )
    qr = qrcode.QRCode(version=None,
                       error_correction=qrcode.constants.ERROR_CORRECT_L,
                       box_size=3, border=1)
    qr.add_data(qr_text)
    qr.make(fit=True)
    qr_img  = qr.make_image(fill_color="black", back_color="white").convert("RGB")
    qr_size = min(QR_W - 4, ROW_H * QR_ROWS - 4)
    qr_img  = qr_img.resize((qr_size, qr_size), Image.LANCZOS)
    img.paste(qr_img, (X_QR + (QR_W - qr_size)//2, CY + (ROW_H*QR_ROWS - qr_size)//2))

    # 8 行資料
    rows_def = [
        ("供應商",        str(record.get("供應商名稱") or "")),
        ("進貨日期",      str(record.get("進貨日期") or "")),
        ("材質/特殊",     f"{record.get('材質','')}{record.get('特殊','')}"),
        ("尺寸",          str(record.get("尺寸") or "")),
        ("長度",          str(record.get("長度") or "")),
        ("批號",          str(record.get("批號") or "")),
        ("數量",          f"{pkg_no}/{pkg_total}"),
        ("製造編號/爐號", str(record.get("製造編號/爐號") or "")),
    ]

    for i, (lbl, val) in enumerate(rows_def):
        y0 = CY + i * ROW_H
        y1 = y0 + ROW_H
        x_right = X_QR if i < QR_ROWS else X_END
        if i > 0:
            draw.line([(CX, y0), (X_END if i >= QR_ROWS else x_right, y0)], fill="black", width=1)
        draw.line([(X_DATA, y0), (X_DATA, y1)], fill="black", width=1)
        _draw_cell(draw, X_LBL, y0, X_DATA, y1, lbl, font_lbl, "center")
        _draw_cell(draw, X_DATA, y0, x_right, y1, val, font_data, "left")

    draw.line([(X_QR, CY), (X_QR, CY + ROW_H * QR_ROWS)], fill="black", width=1)

    # 底部行
    BY = CY + MAIN_H
    BH = BOT_H
    draw.line([(CX, BY),      (X_END, BY)],      fill="black", width=2)
    draw.line([(CX, BY + BH), (X_END, BY + BH)], fill="black", width=2)
    b1 = CX + int(CW * 0.22)
    b2 = CX + int(CW * 0.50)
    b3 = CX + int(CW * 0.72)
    for x in (b1, b2, b3):
        draw.line([(x, BY), (x, BY + BH)], fill="black", width=1)
    _draw_cell(draw, CX, BY, b1,    BY+BH, "ERP序號",                         font_bot, "center")
    _draw_cell(draw, b1,  BY, b2,   BY+BH, str(record.get("序號") or ""),     font_bot, "center")
    _draw_cell(draw, b2,  BY, b3,   BY+BH, "訂單編號",                        font_bot, "center")
    _draw_cell(draw, b3,  BY, X_END,BY+BH, str(record.get("訂單編號") or ""), font_bot, "center")

    # 外框
    draw.rectangle([(CX, CY), (X_END, BY + BH)], outline="black", width=2)
    return img


# ── Windows 列印 ──────────────────────────────────────────────────────────────

def print_label_simple(printer_name, pil_image):
    """強制以 LABEL_W_MM×LABEL_H_MM 為單張紙尺寸列印"""
    DM_ORIENTATION = 0x00000001
    DM_PAPERSIZE   = 0x00000002
    DM_PAPERLENGTH = 0x00000004
    DM_PAPERWIDTH  = 0x00000008
    DMPAPER_USER   = 256

    hprinter = win32print.OpenPrinter(printer_name)
    try:
        props   = win32print.GetPrinter(hprinter, 2)
        devmode = props["pDevMode"]
        devmode.PaperSize   = DMPAPER_USER
        devmode.PaperWidth  = int(round(LABEL_W_MM * 10))
        devmode.PaperLength = int(round(LABEL_H_MM * 10))
        devmode.Orientation = 1
        devmode.Fields = (devmode.Fields
                          | DM_PAPERSIZE | DM_PAPERWIDTH
                          | DM_PAPERLENGTH | DM_ORIENTATION)
        hdc_int = win32gui.CreateDC("WINSPOOL", printer_name, devmode)
        hdc = win32ui.CreateDCFromHandle(hdc_int)
    finally:
        win32print.ClosePrinter(hprinter)

    dpi_x  = hdc.GetDeviceCaps(win32con.LOGPIXELSX)
    dpi_y  = hdc.GetDeviceCaps(win32con.LOGPIXELSY)
    page_w = hdc.GetDeviceCaps(win32con.HORZRES)
    page_h = hdc.GetDeviceCaps(win32con.VERTRES)
    px_w = int(LABEL_W_MM / 25.4 * dpi_x)
    px_h = int(LABEL_H_MM / 25.4 * dpi_y)
    if page_w and page_h:
        px_w = min(px_w, page_w)
        px_h = min(px_h, page_h)

    img_resized = pil_image.resize((px_w, px_h), Image.LANCZOS).convert("RGB")
    hdc.StartDoc("進料標籤")
    hdc.StartPage()
    from PIL.ImageWin import Dib
    Dib(img_resized).draw(hdc.GetHandleOutput(), (0, 0, px_w, px_h))
    hdc.EndPage()
    hdc.EndDoc()
    hdc.DeleteDC()


# ── 主視窗 ────────────────────────────────────────────────────────────────────

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(APP_TITLE)
        self._selected_ids = set()
        self._cfg = load_config()
        self._build_ui()
        self._restore_window()
        self._refresh_table()
        self.protocol("WM_DELETE_WINDOW", self._on_close)

    # ── 視窗記憶 ──────────────────────────────────────────────────────────────

    def _restore_window(self):
        geo = self._cfg.get("geometry", "1200x700")
        try:
            self.geometry(geo)
        except Exception:
            self.geometry("1200x700")
        saved_printer = self._cfg.get("printer", "")
        if saved_printer and saved_printer in self._printer_cb["values"]:
            self._printer_var.set(saved_printer)

    def _on_close(self):
        self._cfg["geometry"] = self.geometry()
        self._cfg["printer"]  = self._printer_var.get()
        save_config(self._cfg)
        self.destroy()

    # ── UI 建構 ───────────────────────────────────────────────────────────────

    def _build_ui(self):
        # 工具列
        toolbar = ttk.Frame(self, padding=4)
        toolbar.pack(fill="x", side="top")

        ttk.Button(toolbar, text="📂 匯入 Excel", command=self._import_excel).pack(side="left", padx=2)
        ttk.Separator(toolbar, orient="vertical").pack(side="left", padx=6, fill="y")

        ttk.Label(toolbar, text="印表機:").pack(side="left")
        self._printer_var = tk.StringVar()
        self._printer_cb  = ttk.Combobox(toolbar, textvariable=self._printer_var, width=30, state="readonly")
        self._printer_cb.pack(side="left", padx=4)
        self._refresh_printers()
        ttk.Button(toolbar, text="🔄", command=self._refresh_printers, width=3).pack(side="left")

        ttk.Separator(toolbar, orient="vertical").pack(side="left", padx=6, fill="y")
        ttk.Button(toolbar, text="☑ 全選",    command=self._select_all).pack(side="left", padx=2)
        ttk.Button(toolbar, text="☐ 取消全選", command=self._deselect_all).pack(side="left", padx=2)

        ttk.Separator(toolbar, orient="vertical").pack(side="left", padx=6, fill="y")
        ttk.Button(toolbar, text="🔍 預覽標籤", command=self._preview_label).pack(side="left", padx=2)
        ttk.Button(toolbar, text="🖨 列印選取", command=self._print_selected).pack(side="left", padx=2)
        ttk.Button(toolbar, text="🔢 指定張數", command=self._print_range).pack(side="left", padx=2)

        ttk.Separator(toolbar, orient="vertical").pack(side="left", padx=6, fill="y")
        ttk.Button(toolbar, text="🗑 清除資料", command=self._clear_data).pack(side="left", padx=2)

        # 資料表格
        frame = ttk.Frame(self)
        frame.pack(fill="both", expand=True, padx=6, pady=4)

        cols = ("chk", "已印") + tuple(c[0] for c in COLUMNS) + ("每包數量",)
        self._tree = ttk.Treeview(frame, columns=cols, show="headings", selectmode="extended")

        self._tree.heading("chk", text="✓")
        self._tree.column("chk", width=30, anchor="center", stretch=False)

        self._tree.heading("已印", text="已印")
        self._tree.column("已印", width=42, anchor="center", stretch=False)

        for name, w in COLUMNS:
            self._tree.heading(name, text=name)
            self._tree.column(name, width=w, anchor="center")
        self._tree.heading("每包數量", text="每包數量")
        self._tree.column("每包數量", width=70, anchor="center")

        # 已印標記樣式：綠色背景
        self._tree.tag_configure("printed", background="#d6f0d6")

        vsb = ttk.Scrollbar(frame, orient="vertical",   command=self._tree.yview)
        hsb = ttk.Scrollbar(frame, orient="horizontal",  command=self._tree.xview)
        self._tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self._tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        frame.rowconfigure(0, weight=1)
        frame.columnconfigure(0, weight=1)

        self._tree.bind("<ButtonRelease-1>", self._on_tree_click)
        self._tree.bind("<Button-3>",        self._on_right_click)

        # 狀態列
        self._status = tk.StringVar(value="就緒")
        ttk.Label(self, textvariable=self._status, relief="sunken", anchor="w").pack(
            fill="x", side="bottom", padx=4, pady=2)

    # ── 輔助方法 ──────────────────────────────────────────────────────────────

    def _refresh_printers(self):
        printers = [p[2] for p in win32print.EnumPrinters(
            win32print.PRINTER_ENUM_LOCAL | win32print.PRINTER_ENUM_CONNECTIONS)]
        self._printer_cb["values"] = printers
        default = win32print.GetDefaultPrinter()
        if default in printers:
            self._printer_var.set(default)
        elif printers:
            self._printer_var.set(printers[0])

    def _refresh_table(self):
        self._tree.delete(*self._tree.get_children())
        rows    = load_table_data()
        printed = get_printed_sns()

        for r in rows:
            sn  = r["SN"] if "SN" in r.keys() else ""
            pkg = r["PKGQTY"] if r["PKGQTY"] else 1
            is_printed = int(sn) in printed if sn else False
            vals = (
                "☐",
                "✓" if is_printed else "",
                r["序號"] or "", r["供應商名稱"] or "", r["訂單編號"] or "",
                r["材質"] or "", r["尺寸"] or "", r["批號"] or "", r["特殊"] or "",
                r["長度"] or "", r["數量"] or "", r["製造編號/爐號"] or "", r["進貨日期"] or "",
                pkg,
            )
            tags = ("printed",) if is_printed else ()
            self._tree.insert("", "end", iid=str(sn), values=vals, tags=tags)

        cnt = len(rows)
        p   = len(printed)
        self._status.set(f"共 {cnt} 筆資料，已印 {p} 筆，未印 {cnt - p} 筆")

    def _on_tree_click(self, event):
        region = self._tree.identify("region", event.x, event.y)
        col    = self._tree.identify_column(event.x)
        if region == "cell" and col == "#1":
            item = self._tree.identify_row(event.y)
            if item:
                self._toggle_check(item)

    def _toggle_check(self, item):
        vals = list(self._tree.item(item, "values"))
        if vals[0] == "☐":
            vals[0] = "☑"
            self._selected_ids.add(item)
        else:
            vals[0] = "☐"
            self._selected_ids.discard(item)
        self._tree.item(item, values=vals)
        self._status.set(f"已選取 {len(self._selected_ids)} 筆")

    def _select_all(self):
        self._selected_ids.clear()
        for item in self._tree.get_children():
            vals = list(self._tree.item(item, "values"))
            vals[0] = "☑"
            self._tree.item(item, values=vals)
            self._selected_ids.add(item)
        self._status.set(f"已全選 {len(self._selected_ids)} 筆")

    def _deselect_all(self):
        for item in self._tree.get_children():
            vals = list(self._tree.item(item, "values"))
            vals[0] = "☐"
            self._tree.item(item, values=vals)
        self._selected_ids.clear()
        self._status.set("已取消全選")

    # ── 右鍵選單 ──────────────────────────────────────────────────────────────

    def _on_right_click(self, event):
        item = self._tree.identify_row(event.y)
        if not item:
            return
        # 右鍵同時選取該列（若尚未勾選，自動勾選）
        if item not in self._selected_ids:
            self._toggle_check(item)

        menu = tk.Menu(self, tearoff=0)
        menu.add_command(label="🔍 預覽標籤",   command=lambda: self._preview_single(item))
        menu.add_command(label="🖨  列印此筆",   command=lambda: self._print_single(item))
        menu.add_command(label="🔢 指定張數",   command=lambda: self._print_range_for(item))
        menu.add_separator()
        menu.add_command(label="📋 列印記錄",   command=lambda: self._show_print_log(item))
        menu.add_separator()
        menu.add_command(label="🗑 刪除此筆",   command=lambda: self._delete_single(item))
        try:
            menu.tk_popup(event.x_root, event.y_root)
        finally:
            menu.grab_release()

    def _preview_single(self, item):
        row = self._get_record_by_sn(item)
        if not row:
            return
        rec = self._build_record_dict(row)
        img = make_label_image(rec, 1, int(rec.get("數量") or 1))
        PreviewWindow(self, img)

    def _print_single(self, item):
        printer = self._printer_var.get()
        if not printer:
            messagebox.showwarning("提示", "請選擇印表機")
            return
        row = self._get_record_by_sn(item)
        if not row:
            return
        rec       = self._build_record_dict(row)
        total_qty = int(rec.get("數量") or 1)
        if not messagebox.askyesno("確認列印",
                f"列印此筆 {total_qty} 張\n印表機：{printer}\n確定？"):
            return
        jobs = [(rec, i, total_qty, int(item)) for i in range(1, total_qty + 1)]
        PrintJobDialog(self, jobs, printer, callback=self._refresh_table)

    def _print_range_for(self, item):
        row = self._get_record_by_sn(item)
        if not row:
            return
        rec     = self._build_record_dict(row)
        total   = int(rec.get("數量") or 1)
        printer = self._printer_var.get()
        if not printer:
            messagebox.showwarning("提示", "請選擇印表機")
            return
        PrintRangeDialog(self, rec, total, printer)

    def _show_print_log(self, item):
        """顯示列印記錄視窗"""
        logs = get_print_history(int(item))
        win  = tk.Toplevel(self)
        win.title(f"列印記錄 — SN {item}")
        win.geometry("520x340")
        win.transient(self)
        win.grab_set()

        cols = ("時間", "印表機", "張數")
        tv   = ttk.Treeview(win, columns=cols, show="headings", height=10)
        tv.heading("時間",  text="列印時間")
        tv.heading("印表機", text="印表機")
        tv.heading("張數",  text="列印張數")
        tv.column("時間",  width=160, anchor="center")
        tv.column("印表機", width=200, anchor="w")
        tv.column("張數",  width=80,  anchor="center")

        if not logs:
            tv.insert("", "end", values=("（尚無記錄）", "", ""))
        else:
            for log in logs:
                at, printer, nums_json = log
                try:
                    nums = json.loads(nums_json) if nums_json else []
                    cnt  = len(nums)
                except Exception:
                    cnt = "?"
                tv.insert("", "end", values=(at, printer, cnt))

        tv.pack(fill="both", expand=True, padx=10, pady=10)
        ttk.Button(win, text="關閉", command=win.destroy).pack(pady=6)

    def _delete_single(self, item):
        if not messagebox.askyesno("確認刪除", f"確定刪除 SN={item} 這筆資料？\n此動作無法還原"):
            return
        conn = get_db()
        conn.execute("DELETE FROM MYTABLE WHERE SN=?", (int(item),))
        conn.commit()
        conn.close()
        self._selected_ids.discard(item)
        self._refresh_table()

    # ── 功能方法 ──────────────────────────────────────────────────────────────

    def _import_excel(self):
        path = filedialog.askopenfilename(
            title="選擇 ERP Excel 檔案",
            filetypes=[("Excel 檔案", "*.xlsx *.xls"), ("所有檔案", "*.*")]
        )
        if not path:
            return
        try:
            cnt = import_excel_to_db(path)
            self._refresh_table()
            messagebox.showinfo("匯入成功", f"成功匯入 {cnt} 筆資料")
        except Exception as e:
            messagebox.showerror("匯入失敗", str(e))

    def _get_record_by_sn(self, sn):
        conn = get_db()
        cur  = conn.cursor()
        cur.execute("SELECT * FROM MYTABLE WHERE SN=?", (sn,))
        row  = cur.fetchone()
        conn.close()
        return row

    def _build_record_dict(self, row):
        d = dict(row)
        if "製造編號/爐號" not in d:
            d["製造編號/爐號"] = d.get("製造編號爐號", "")
        return d

    def _preview_label(self):
        if not self._selected_ids:
            messagebox.showwarning("提示", "請先勾選要預覽的資料")
            return
        sn  = next(iter(self._selected_ids))
        row = self._get_record_by_sn(sn)
        if not row:
            return
        rec = self._build_record_dict(row)
        img = make_label_image(rec, 1, int(rec.get("數量") or 1))
        PreviewWindow(self, img)

    def _print_selected(self):
        if not self._selected_ids:
            messagebox.showwarning("提示", "請先勾選要列印的資料")
            return
        printer = self._printer_var.get()
        if not printer:
            messagebox.showwarning("提示", "請選擇印表機")
            return

        # 統計總張數
        jobs = []
        for sn in self._selected_ids:
            row = self._get_record_by_sn(sn)
            if not row:
                continue
            rec       = self._build_record_dict(row)
            total_qty = int(rec.get("數量") or 1)
            for i in range(1, total_qty + 1):
                jobs.append((rec, i, total_qty, int(sn)))

        if not jobs:
            return

        total_labels = len(jobs)
        if not messagebox.askyesno("確認列印",
                f"即將列印 {len(self._selected_ids)} 筆，共 {total_labels} 張標籤\n"
                f"印表機：{printer}\n\n確定列印？"):
            return

        PrintJobDialog(self, jobs, printer, callback=self._refresh_table)

    def _parse_range(self, text, max_val):
        nums = set()
        for part in text.replace('，', ',').replace('~', '-').replace('到', '-').split(','):
            part = part.strip()
            if not part:
                continue
            if '-' in part:
                a, _, b = part.partition('-')
                try:
                    nums.update(range(int(a.strip()), int(b.strip()) + 1))
                except ValueError:
                    pass
            else:
                try:
                    nums.add(int(part))
                except ValueError:
                    pass
        return sorted(n for n in nums if 1 <= n <= max_val)

    def _print_range(self):
        if not self._selected_ids:
            messagebox.showwarning("提示", "請先勾選一筆要重印的資料")
            return
        if len(self._selected_ids) > 1:
            messagebox.showwarning("提示", "指定張數每次只能選一筆資料")
            return
        sn  = next(iter(self._selected_ids))
        row = self._get_record_by_sn(sn)
        if not row:
            return
        rec     = self._build_record_dict(row)
        total   = int(rec.get("數量") or 1)
        printer = self._printer_var.get()
        if not printer:
            messagebox.showwarning("提示", "請選擇印表機")
            return
        PrintRangeDialog(self, rec, total, printer)

    def _clear_data(self):
        if not messagebox.askyesno("確認", "確定要清除所有資料嗎？\n此動作無法還原"):
            return
        conn = get_db()
        conn.execute("DELETE FROM MYTABLE")
        conn.commit()
        conn.close()
        self._selected_ids.clear()
        self._refresh_table()


# ── 批次列印進度對話框 ────────────────────────────────────────────────────────

class PrintJobDialog(tk.Toplevel):
    """多頁批次列印，顯示進度條與取消按鈕"""

    def __init__(self, parent, jobs, printer, callback=None):
        """
        jobs     : list of (rec, label_no, label_total, sn)
        callback : 列印完成後呼叫（通常是 _refresh_table）
        """
        super().__init__(parent)
        self.title("列印中...")
        self._jobs     = jobs
        self._printer  = printer
        self._callback = callback
        self._parent   = parent
        self._idx      = 0
        self._success  = 0
        self._cancelled = False
        self._printed_log: dict[int, list] = {}  # sn -> [label_nos]

        total = len(jobs)
        w, h = 460, 200
        px = parent.winfo_rootx() + parent.winfo_width()  // 2 - w // 2
        py = parent.winfo_rooty() + parent.winfo_height() // 2 - h // 2
        self.geometry(f"{w}x{h}+{max(0,px)}+{max(0,py)}")
        self.resizable(False, False)
        self.transient(parent)
        self.protocol("WM_DELETE_WINDOW", self._cancel)

        # 目前狀態文字
        self._lbl_info = ttk.Label(self, text="準備中...", anchor="w")
        self._lbl_info.pack(fill="x", padx=16, pady=(20, 4))

        # 進度條
        self._bar = ttk.Progressbar(self, maximum=total, value=0, length=420)
        self._bar.pack(padx=16, pady=4)

        # 數字進度
        self._lbl_count = ttk.Label(self, text=f"0 / {total}", foreground="#555")
        self._lbl_count.pack(pady=2)

        # 取消按鈕
        self._btn_cancel = ttk.Button(self, text="取消", width=12, command=self._cancel)
        self._btn_cancel.pack(pady=12)

        self.grab_set()
        self.after(80, self._step)

    def _step(self):
        if self._cancelled or self._idx >= len(self._jobs):
            self._finish()
            return

        rec, n, total_qty, sn = self._jobs[self._idx]
        sup = rec.get("供應商名稱", "")
        self._lbl_info.config(
            text=f"列印第 {self._idx+1} 張  供應商：{sup}  張數：{n}/{total_qty}")
        self.update_idletasks()

        try:
            img = make_label_image(rec, n, total_qty)
            print_label_simple(self._printer, img)
            self._success += 1
            # 記錄此 SN 的已印張號
            self._printed_log.setdefault(sn, []).append(n)
        except Exception as e:
            messagebox.showerror("列印錯誤",
                f"第 {n} 張（{sup}）失敗：{e}", parent=self)

        self._idx += 1
        self._bar["value"] = self._idx
        self._lbl_count.config(text=f"{self._idx} / {len(self._jobs)}")
        self.after(10, self._step)

    def _cancel(self):
        self._cancelled = True

    def _finish(self):
        # 寫入列印記錄
        for sn, nums in self._printed_log.items():
            log_print(sn, self._printer, nums)

        status = "已取消" if self._cancelled else "完成"
        self._lbl_info.config(
            text=f"列印{status}！成功 {self._success} 張，共 {len(self._jobs)} 張")
        self._bar["value"] = self._idx
        self._btn_cancel.config(text="關閉", command=self.destroy)

        if not self._cancelled:
            self.after(1200, self.destroy)  # 完成後自動關閉

        if self._callback:
            self.after(100, self._callback)


# ── 指定張數對話框（含即時預覽）────────────────────────────────────────────────

class PrintRangeDialog(tk.Toplevel):
    """指定張數列印對話框 — 左側輸入 + 右側即時標籤預覽"""

    def __init__(self, parent, rec, total, printer):
        super().__init__(parent)
        self.title("指定張數列印")
        self._rec     = rec
        self._total   = total
        self._printer = printer
        self._parent  = parent

        w, h = 860, 400
        px = parent.winfo_rootx() + parent.winfo_width()  // 2 - w // 2
        py = parent.winfo_rooty() + parent.winfo_height() // 2 - h // 2
        self.geometry(f"{w}x{h}+{max(0,px)}+{max(0,py)}")
        self.minsize(w, h)
        self.resizable(False, False)

        self._build(rec, total)
        self.update_idletasks()
        self.transient(parent)
        self.lift()
        self.focus_force()
        self.after(50, self.grab_set)
        # 初始預覽第 1 張
        self.after(100, lambda: self._update_preview(1))

    def _build(self, rec, total):
        # ── 左右兩欄 ──────────────────────────────────────────────────────────
        left  = ttk.Frame(self, width=480)
        right = ttk.Frame(self, width=360, relief="groove", borderwidth=1)
        left.pack(side="left", fill="both", expand=False, padx=(10,4), pady=10)
        right.pack(side="left", fill="both", expand=True, padx=(4,10), pady=10)
        left.pack_propagate(False)

        # 左側：輸入控制
        info = (f"供應商：{rec.get('供應商名稱','')}　"
                f"批號：{rec.get('批號','')}　"
                f"總數量：{total} 張")
        ttk.Label(left, text=info, foreground="#444",
                  wraplength=450).pack(padx=6, pady=(8,2), anchor="w")
        ttk.Separator(left, orient="horizontal").pack(fill="x", padx=6, pady=4)

        ttk.Label(left, text="輸入要列印的張數（支援逗號與範圍）：").pack(padx=6, anchor="w")
        ttk.Label(left, text="範例：  1,2,3,5   或   30-33   或   1-3,5,30-33",
                  foreground="gray").pack(padx=6, anchor="w")

        self._entry = ttk.Entry(left, width=36, font=("微軟正黑體", 12))
        self._entry.pack(padx=6, pady=6, fill="x")
        self._entry.focus()

        # 快捷按鈕
        qf = ttk.Frame(left)
        qf.pack(padx=6, pady=2, fill="x")
        ttk.Label(qf, text="快速：").pack(side="left")
        ttk.Button(qf, text=f"全部 1-{total}",
                   command=lambda: self._set(f"1-{total}")).pack(side="left", padx=2)
        ttk.Button(qf, text="最後5張",
                   command=lambda: self._set(f"{max(1,total-4)}-{total}")).pack(side="left", padx=2)
        ttk.Button(qf, text="第1張",
                   command=lambda: self._set("1")).pack(side="left", padx=2)

        # 結果提示
        self._preview_lbl = ttk.Label(left, text="", foreground="#0066cc")
        self._preview_lbl.pack(padx=6, pady=4, anchor="w")
        self._entry.bind("<KeyRelease>", self._on_key)

        ttk.Separator(left, orient="horizontal").pack(fill="x", padx=6, pady=6)

        bf = ttk.Frame(left)
        bf.pack(padx=6, pady=(0,8))
        ttk.Button(bf, text="列印", width=12, command=self._do_print).pack(side="left", padx=4)
        ttk.Button(bf, text="取消", width=12, command=self.destroy).pack(side="left", padx=4)

        # 右側：即時預覽
        ttk.Label(right, text="標籤預覽", foreground="#666").pack(pady=(8,2))
        ttk.Separator(right, orient="horizontal").pack(fill="x", padx=4)
        self._canvas = tk.Canvas(right, bg="#f0f0f0", cursor="arrow")
        self._canvas.pack(fill="both", expand=True, padx=6, pady=6)
        self._preview_photo = None

    def _update_preview(self, label_no=None):
        """在右側畫布顯示指定張號的標籤預覽"""
        if label_no is None:
            nums = self._parent._parse_range(self._entry.get(), self._total)
            label_no = nums[0] if nums else 1

        try:
            img = make_label_image(self._rec, label_no, self._total)
            # 依畫布大小縮放
            cw = self._canvas.winfo_width()  or 340
            ch = self._canvas.winfo_height() or 340
            scale = min((cw - 10) / img.width, (ch - 10) / img.height)
            nw = max(1, int(img.width  * scale))
            nh = max(1, int(img.height * scale))
            preview = img.resize((nw, nh), Image.LANCZOS)
            from PIL import ImageTk
            self._preview_photo = ImageTk.PhotoImage(preview)
            self._canvas.delete("all")
            self._canvas.create_image(cw//2, ch//2, anchor="center",
                                      image=self._preview_photo)
        except Exception:
            pass

    def _set(self, text):
        self._entry.delete(0, tk.END)
        self._entry.insert(0, text)
        self._on_key()

    def _on_key(self, event=None):
        nums = self._parent._parse_range(self._entry.get(), self._total)
        if nums:
            self._preview_lbl.config(
                text=f"將列印 {len(nums)} 張：{', '.join(map(str, nums[:10]))}{'...' if len(nums)>10 else ''}",
                foreground="#0066cc")
            self._update_preview(nums[0])
        else:
            self._preview_lbl.config(text="（尚未輸入有效張數）", foreground="gray")

    def _do_print(self):
        nums = self._parent._parse_range(self._entry.get(), self._total)
        if not nums:
            messagebox.showwarning("提示", "請輸入有效的張數", parent=self)
            return
        if not messagebox.askyesno("確認列印",
                f"即將列印 {len(nums)} 張\n印表機：{self._printer}\n\n確定？", parent=self):
            return

        sn   = None
        for k, v in self._rec.items():
            if k == "SN":
                sn = v
                break
        if sn is None:
            sn = 0

        jobs = [(self._rec, n, self._total, int(sn)) for n in nums]
        self.destroy()
        PrintJobDialog(self._parent, jobs, self._printer,
                       callback=self._parent._refresh_table)


# ── 預覽視窗 ──────────────────────────────────────────────────────────────────

class PreviewWindow(tk.Toplevel):
    def __init__(self, parent, pil_image):
        super().__init__(parent)
        self.title("標籤預覽")
        self.resizable(False, False)

        scale   = 0.8
        w = int(pil_image.width  * scale)
        h = int(pil_image.height * scale)
        preview = pil_image.resize((w, h), Image.LANCZOS)

        from PIL import ImageTk
        self._photo = ImageTk.PhotoImage(preview)
        tk.Label(self, image=self._photo, relief="solid", bd=1).pack(padx=10, pady=10)
        ttk.Label(self,
                  text=f"實際尺寸：{LABEL_W_MM}mm × {LABEL_H_MM}mm  /  "
                       f"{LABEL_W_PX}px × {LABEL_H_PX}px @ {PRINT_DPI}DPI",
                  foreground="gray").pack(pady=2)
        ttk.Button(self, text="關閉", command=self.destroy).pack(pady=6)
        self.transient(parent)
        self.grab_set()


# ── 啟動 ──────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    init_db()
    app = App()
    app.mainloop()
